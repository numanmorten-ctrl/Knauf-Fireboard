from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

KNAUF_BLUE = "009FE3"
WHITE = "FFFFFF"
DEFAULT_SHEET_NAME = "Materialeliste"
LOGO_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "PDF_template.pdf"
LOGO_ANCHOR = "A1"
TITLE_ROW = 4
TABLE_START_ROW = 6


def add_system_separator_rows(materials_df, system_column="SYSTEM"):
    """
    Sort materials by system and add plain grouped export rows.
    """

    if system_column not in materials_df.columns:

        return materials_df.copy()

    export_columns = [
        system_column,
        *(
            column
            for column in materials_df.columns
            if column != system_column
        )
    ]

    sorted_df = (
        materials_df[export_columns]
        .sort_values(
            by=system_column,
            kind="stable"
        )
    )

    rows = []
    previous_system = object()

    for _, row in sorted_df.iterrows():

        current_system = row[system_column]
        current_system_key = (
            ("__MISSING_SYSTEM__",)
            if pd.isna(current_system)
            else current_system
        )

        if current_system_key != previous_system:

            if rows:

                rows.append({
                    column: ""
                    for column in export_columns
                })

            rows.append(dict(zip(export_columns, export_columns)))

            system_row = {
                column: ""
                for column in export_columns
            }
            system_row[system_column] = current_system
            rows.append(system_row)

            previous_system = current_system_key

        rows.append(row.to_dict())

    return pd.DataFrame(
        rows,
        columns=export_columns
    )


def _get_material_list_title(language="DA", per_system=False):
    """Return the localized Excel title without affecting export data."""

    if language == "EN":
        return "Material list per system" if per_system else "Material list"

    return "Materialeliste pr. system" if per_system else "Materialeliste"


def _extract_logo_from_pdf_template(template_path=LOGO_TEMPLATE_PATH):
    """Load the existing Knauf logo image embedded in the repository PDF template."""

    from pypdf import PdfReader

    pdf_reader = PdfReader(str(template_path))

    for page in pdf_reader.pages:

        for image in page.images:

            if image.data:

                return BytesIO(image.data)

    raise FileNotFoundError(
        f"No embedded logo image found in {template_path}"
    )


def _add_knauf_logo(worksheet, template_path=LOGO_TEMPLATE_PATH):
    """Insert the existing repository Knauf logo above the worksheet table."""

    from openpyxl.drawing.image import Image as OpenpyxlImage

    logo_stream = _extract_logo_from_pdf_template(template_path)
    logo = OpenpyxlImage(logo_stream)
    logo.width = 170
    logo.height = 54
    worksheet.add_image(logo, LOGO_ANCHOR)


def _row_values(worksheet, row_number):
    return [
        worksheet.cell(row=row_number, column=column_number).value
        for column_number in range(1, worksheet.max_column + 1)
    ]


def _is_blank(value):
    return value is None or str(value).strip() == ""


def _is_header_row(values, columns):
    return [str(value) for value in values[:len(columns)]] == [
        str(column)
        for column in columns
    ]


def _is_system_name_row(values, columns):
    if not columns or str(columns[0]) != "SYSTEM":
        return False

    return (
        not _is_blank(values[0])
        and all(_is_blank(value) for value in values[1:])
        and not _is_header_row(values, columns)
    )


def _style_header_row(worksheet, row_number):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=KNAUF_BLUE
    )
    header_font = Font(
        color=WHITE,
        bold=True
    )

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font


def _style_system_name_row(worksheet, row_number):
    for cell in worksheet[row_number]:
        cell.font = Font(bold=True)


def _autosize_columns(worksheet):
    for column_number in range(1, worksheet.max_column + 1):
        max_length = 0

        for row_number in range(1, worksheet.max_row + 1):
            value = worksheet.cell(
                row=row_number,
                column=column_number
            ).value

            if value is not None:
                max_length = max(max_length, len(str(value)))

        adjusted_width = min(max(max_length + 4, 12), 60)
        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = adjusted_width


def _apply_materials_excel_styling(
    worksheet,
    columns,
    language="DA",
    per_system=False,
    include_header=True,
    table_start_row=TABLE_START_ROW
):
    _add_knauf_logo(worksheet)

    title_cell = worksheet.cell(
        row=TITLE_ROW,
        column=1,
        value=_get_material_list_title(language, per_system)
    )
    title_cell.font = Font(
        bold=True,
        size=16
    )

    header_rows = []

    if include_header:
        header_rows.append(table_start_row)

    else:
        for row_number in range(table_start_row, worksheet.max_row + 1):
            if _is_header_row(_row_values(worksheet, row_number), columns):
                header_rows.append(row_number)

    for row_number in header_rows:
        _style_header_row(worksheet, row_number)

    if per_system:
        for row_number in range(table_start_row, worksheet.max_row + 1):
            values = _row_values(worksheet, row_number)

            if _is_system_name_row(values, columns):
                _style_system_name_row(worksheet, row_number)

    first_header_row = header_rows[0] if header_rows else table_start_row
    worksheet.freeze_panes = f"A{first_header_row + 1}"
    _autosize_columns(worksheet)


def create_materials_excel(
    materials_df,
    autosize_columns=True,
    include_header=True,
    language="DA",
    per_system=False
):
    """
    Create styled Excel file from materials dataframe.
    """

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        materials_df.to_excel(
            writer,
            index=False,
            header=include_header,
            sheet_name=DEFAULT_SHEET_NAME,
            startrow=TABLE_START_ROW - 1
        )

        worksheet = writer.sheets[DEFAULT_SHEET_NAME]

        _apply_materials_excel_styling(
            worksheet,
            list(materials_df.columns),
            language=language,
            per_system=per_system,
            include_header=include_header
        )

    output.seek(0)

    return output
