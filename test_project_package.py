from io import BytesIO
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from translations import translations
from utils.project_package import (
    build_project_download_signature,
    CUSTOM_PROFILE_MATERIAL_NOTE,
    build_project_material_exports,
    collect_project_external_files,
    create_project_package_zip,
    get_cached_project_download,
)


class FakeResponse:
    def __init__(self, content=b"pdf"):
        self.content = content

    def raise_for_status(self):
        return None


def t_da(key):
    return translations["DA"].get(key, key)


def project_materials():
    return {
        "1. system": pd.DataFrame(
            [
                {
                    "ART.NR.": "2906",
                    "DB NR": "5959671",
                    "PRODUCENT": "Knauf A/S",
                    "BESKRIVELSE": "15 mm Fireboard 1250x2000",
                    "FORBRUG PR. LBM": "1,5",
                    "ENHED": "m²",
                    "SAMLET MÆNGDE": "10,0",
                },
                {
                    "ART.NR.": "2907",
                    "DB NR": "5959673",
                    "PRODUCENT": "Knauf A/S",
                    "BESKRIVELSE": "25 mm Fireboard 1250x2000",
                    "FORBRUG PR. LBM": "2,0",
                    "ENHED": "m²",
                    "SAMLET MÆNGDE": "5,0",
                },
            ]
        ),
        "2. system": pd.DataFrame(
            [
                {
                    "ART.NR.": "2906",
                    "DB NR": "5959671",
                    "PRODUCENT": "Knauf A/S",
                    "BESKRIVELSE": "15 mm Fireboard 1250x2000",
                    "FORBRUG PR. LBM": "1,0",
                    "ENHED": "m²",
                    "SAMLET MÆNGDE": "7,0",
                }
            ]
        ),
    }



def workbook_values(excel_file):
    excel_file.seek(0)
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    return [
        [cell.value for cell in row]
        for row in worksheet.iter_rows()
    ]


def flattened_workbook_text(excel_file):
    return "\n".join(
        str(value)
        for row in workbook_values(excel_file)
        for value in row
        if value is not None
    )


def custom_calculation(profile="Special profil"):
    return {
        "category": "Andre profiler",
        "profile": profile,
        "fire_time": 60,
        "temperature": 350,
        "custom_apv": 123,
    }


def standard_calculation(profile="HEB 100"):
    return {
        "category": "Bjælker",
        "profile": profile,
        "fire_time": 60,
        "temperature": 350,
    }

def material_lookup_records():
    return [
        {
            "ART.NR.": "2906",
            "DB_NR.": "5959671",
            "BESKRIVELSE_DK": "15 mm Fireboard 1250x2000",
            "BESKRIVELSE_EN": "15 mm Fireboard 1250x2000",
            "EPD_URL": "https://example.test/shared-epd.pdf",
            "DATABLAD_URL": "https://example.test/shared-datasheet.pdf",
        },
        {
            "ART.NR.": "2907",
            "DB_NR.": "5959673",
            "BESKRIVELSE_DK": "25 mm Fireboard 1250x2000",
            "BESKRIVELSE_EN": "25 mm Fireboard 1250x2000",
            "EPD_URL": "https://example.test/shared-epd.pdf",
            "DATABLAD_URL": "https://example.test/shared-datasheet.pdf",
        },
    ]


def test_project_package_zip_is_created_when_calculations_exist():
    material_exports = build_project_material_exports(
        project_materials(),
        "DA",
        t_da,
    )

    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=material_exports,
        combined_materials=project_materials(),
        materials_lookup_records=material_lookup_records(),
        language="DA",
        fetcher=lambda url, timeout=20: FakeResponse(f"content:{url}".encode()),
    )

    with ZipFile(package) as archive:
        assert archive.namelist()
        assert "Rapport/Knauf_Fireboard_Rapport.pdf" in archive.namelist()


def test_project_package_contains_report_and_material_list_exports():
    material_exports = build_project_material_exports(
        project_materials(),
        "DA",
        t_da,
    )

    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=material_exports,
        combined_materials=project_materials(),
        materials_lookup_records=material_lookup_records(),
        language="DA",
        fetcher=lambda url, timeout=20: FakeResponse(b"external"),
    )

    with ZipFile(package) as archive:
        names = set(archive.namelist())
        assert "Rapport/Knauf_Fireboard_Rapport.pdf" in names
        assert "Materialelister/Samlet_materialeliste.xlsx" in names
        assert "Materialelister/Materialeliste_pr_beregning.xlsx" in names
        assert archive.read("Rapport/Knauf_Fireboard_Rapport.pdf") == b"report"


def test_project_package_uses_danish_fireboard_documentation_folder_and_names():
    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=None,
        combined_materials={},
        materials_lookup_records=[],
        language="DA",
        fetcher=lambda url, timeout=20: FakeResponse(b"external"),
    )

    with ZipFile(package) as archive:
        names = set(archive.namelist())
        assert (
            "Projekterings- og montageafsnit/Fireboard_projekteringsafsnit.pdf"
            in names
        )
        assert (
            "Projekterings- og montageafsnit/Fireboard_montageafsnit.pdf"
            in names
        )
        assert not any(name.startswith("Fireboard/") for name in names)


def test_project_package_uses_english_fireboard_documentation_folder_and_names():
    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=None,
        combined_materials={},
        materials_lookup_records=[],
        language="EN",
        fetcher=lambda url, timeout=20: FakeResponse(b"external"),
    )

    with ZipFile(package) as archive:
        names = set(archive.namelist())
        assert (
            "Design and Installation Sections/Fireboard_design_section.pdf"
            in names
        )
        assert (
            "Design and Installation Sections/Fireboard_installation_section.pdf"
            in names
        )
        assert not any(name.startswith("Fireboard/") for name in names)


def test_skipped_fireboard_documentation_readme_uses_updated_label():
    def failing_fetcher(url, timeout=20):
        raise RuntimeError("network unavailable")

    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=None,
        combined_materials={},
        materials_lookup_records=[],
        language="DA",
        fetcher=failing_fetcher,
    )

    with ZipFile(package) as archive:
        readme = archive.read("README.txt").decode("utf-8")
        assert "Fireboard projekteringsafsnit" in readme
        assert "Fireboard manualafsnit" not in readme


def test_duplicate_epd_and_datasheet_urls_are_included_once():
    external_files = collect_project_external_files(
        project_materials(),
        material_lookup_records(),
        language="DA",
    )

    epd_files = [file for file in external_files if file.folder == "EPD"]
    datasheet_files = [file for file in external_files if file.folder == "Datablade"]

    assert len(epd_files) == 1
    assert len(datasheet_files) == 1


def test_fireboard_documents_use_friendly_zip_filenames_without_article_numbers():
    external_files = collect_project_external_files(
        project_materials(),
        material_lookup_records(),
        language="DA",
    )

    archive_paths = {f"{file.folder}/{file.filename}" for file in external_files}

    assert "EPD/EPD Fireboard.pdf" in archive_paths
    assert "Datablade/Datablad Fireboard.pdf" in archive_paths
    assert not any("2906" in path or "2907" in path for path in archive_paths)


def test_document_friendly_names_cover_known_product_families():
    combined_materials = {
        "1. system": pd.DataFrame(
            [
                {"BESKRIVELSE": "15 mm Fireboard 1250x2000"},
                {"BESKRIVELSE": "Fireboard spartelmasse 10 kg"},
                {"BESKRIVELSE": "Bjælkeprofil BJ 8-10 rød 2000 mm"},
                {"BESKRIVELSE": "Vinkelprofil 30/30 4000 mm"},
                {"BESKRIVELSE": "Skrue RA51 500 stk."},
            ]
        )
    }
    lookup_records = [
        {
            "BESKRIVELSE_DK": "15 mm Fireboard 1250x2000",
            "BESKRIVELSE_EN": "15 mm Fireboard 1250x2000",
            "EPD_URL": "https://example.test/fireboard-epd",
            "DATABLAD_URL": "https://example.test/fireboard-datasheet",
        },
        {
            "BESKRIVELSE_DK": "Fireboard spartelmasse 10 kg",
            "BESKRIVELSE_EN": "Fireboard joint filler 10 kg",
            "EPD_URL": "https://example.test/spartel-epd",
            "DATABLAD_URL": "https://example.test/spartel-datasheet",
        },
        {
            "BESKRIVELSE_DK": "Bjælkeprofil BJ 8-10 rød 2000 mm",
            "BESKRIVELSE_EN": "Bjælkeprofil BJ 8-10 red 2000 mm",
            "EPD_URL": "",
            "DATABLAD_URL": "https://example.test/bjælkeprofil-datasheet",
        },
        {
            "BESKRIVELSE_DK": "Vinkelprofil 30/30 4000 mm",
            "BESKRIVELSE_EN": "Angle profile 30/30 4000 mm",
            "EPD_URL": "",
            "DATABLAD_URL": "https://example.test/vinkelprofil-datasheet",
        },
        {
            "BESKRIVELSE_DK": "Skrue RA51 500 stk.",
            "BESKRIVELSE_EN": "RA screw 51 500 pcs.",
            "EPD_URL": "",
            "DATABLAD_URL": "https://example.test/ra-screw-datasheet",
        },
    ]

    external_files = collect_project_external_files(
        combined_materials,
        lookup_records,
        language="DA",
    )
    archive_paths = {f"{file.folder}/{file.filename}" for file in external_files}

    assert "EPD/EPD Fireboard.pdf" in archive_paths
    assert "EPD/EPD Fireboard Spartelmasse.pdf" in archive_paths
    assert "Datablade/Datablad Fireboard.pdf" in archive_paths
    assert "Datablade/Datablad Fireboard Spartelmasse.pdf" in archive_paths
    assert "Datablade/Datablad Bilkeprofil.pdf" in archive_paths
    assert "Datablade/Datablad Vinkelprofil.pdf" in archive_paths
    assert "Datablade/Datablad Skrue RA.pdf" in archive_paths


def test_shared_friendly_documents_are_included_only_once_in_zip():
    material_exports = build_project_material_exports(
        project_materials(),
        "DA",
        t_da,
    )

    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=material_exports,
        combined_materials=project_materials(),
        materials_lookup_records=material_lookup_records(),
        language="DA",
        fetcher=lambda url, timeout=20: FakeResponse(f"content:{url}".encode()),
    )

    with ZipFile(package) as archive:
        names = archive.namelist()
        assert names.count("EPD/EPD Fireboard.pdf") == 1
        assert names.count("Datablade/Datablad Fireboard.pdf") == 1


def test_missing_external_files_do_not_break_zip_creation():
    material_exports = build_project_material_exports(
        project_materials(),
        "DA",
        t_da,
    )

    def failing_fetcher(url, timeout=20):
        if "shared-epd" in url:
            raise RuntimeError("network unavailable")
        return FakeResponse(b"external")

    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=material_exports,
        combined_materials=project_materials(),
        materials_lookup_records=material_lookup_records(),
        language="DA",
        fetcher=failing_fetcher,
    )

    with ZipFile(package) as archive:
        names = archive.namelist()
        assert "Rapport/Knauf_Fireboard_Rapport.pdf" in names
        assert "README.txt" in names
        assert "shared-epd" in archive.read("README.txt").decode("utf-8")


def test_project_download_signature_changes_only_for_saved_project_inputs():
    base_signature = build_project_download_signature(
        calculations=[{"profile": "HEB 100", "fire_time": 60}],
        combined_materials=project_materials(),
        language="DA",
        project_details={"project_name": "Project A"},
    )
    same_signature = build_project_download_signature(
        calculations=[{"profile": "HEB 100", "fire_time": 60}],
        combined_materials=project_materials(),
        language="DA",
        project_details={"project_name": "Project A"},
    )
    changed_signature = build_project_download_signature(
        calculations=[{"profile": "HEB 100", "fire_time": 90}],
        combined_materials=project_materials(),
        language="DA",
        project_details={"project_name": "Project A"},
    )

    assert same_signature == base_signature
    assert changed_signature != base_signature


def test_cached_project_download_is_returned_only_for_matching_signature():
    signature = build_project_download_signature(
        calculations=[{"profile": "HEB 100", "fire_time": 60}],
        combined_materials=project_materials(),
        language="DA",
    )
    changed_signature = build_project_download_signature(
        calculations=[{"profile": "HEB 100", "fire_time": 90}],
        combined_materials=project_materials(),
        language="DA",
    )
    cached_zip = BytesIO(b"cached")
    cache = {"signature": signature, "data": cached_zip}

    assert get_cached_project_download(cache, signature) is cached_zip
    assert get_cached_project_download(cache, changed_signature) is None
    assert get_cached_project_download({}, signature) is None


def test_material_exports_support_project_with_only_standard_profiles():
    material_exports = build_project_material_exports(
        project_materials(),
        "DA",
        t_da,
        calculations=[standard_calculation()],
    )

    assert material_exports is not None
    text = flattened_workbook_text(material_exports.combined_excel)
    assert "15 mm Fireboard 1250x2000" in text
    assert CUSTOM_PROFILE_MATERIAL_NOTE not in text


def test_material_exports_support_project_with_only_custom_profiles():
    material_exports = build_project_material_exports(
        {},
        "DA",
        t_da,
        calculations=[custom_calculation()],
    )

    assert material_exports is not None
    assert CUSTOM_PROFILE_MATERIAL_NOTE in flattened_workbook_text(
        material_exports.combined_excel
    )
    assert CUSTOM_PROFILE_MATERIAL_NOTE in flattened_workbook_text(
        material_exports.per_calculation_excel
    )

    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=material_exports,
        combined_materials={},
        materials_lookup_records=[],
        language="DA",
        fetcher=lambda url, timeout=20: FakeResponse(b"external"),
    )
    with ZipFile(package) as archive:
        assert "Materialelister/Samlet_materialeliste.xlsx" in archive.namelist()
        assert "Materialelister/Materialeliste_pr_beregning.xlsx" in archive.namelist()


def test_material_exports_support_mixed_standard_and_custom_profiles():
    material_exports = build_project_material_exports(
        project_materials(),
        "DA",
        t_da,
        calculations=[standard_calculation(), custom_calculation()],
    )

    assert material_exports is not None
    combined_text = flattened_workbook_text(material_exports.combined_excel)
    per_calculation_text = flattened_workbook_text(
        material_exports.per_calculation_excel
    )
    assert "15 mm Fireboard 1250x2000" in combined_text
    assert CUSTOM_PROFILE_MATERIAL_NOTE in combined_text
    assert CUSTOM_PROFILE_MATERIAL_NOTE in per_calculation_text


def test_project_package_allows_missing_material_exports():
    package = create_project_package_zip(
        report_pdf=BytesIO(b"report"),
        material_exports=None,
        combined_materials={},
        materials_lookup_records=[],
        language="DA",
        fetcher=lambda url, timeout=20: FakeResponse(b"external"),
    )

    with ZipFile(package) as archive:
        assert "Rapport/Knauf_Fireboard_Rapport.pdf" in archive.namelist()
        assert not any(
            name.startswith("Materialelister/") for name in archive.namelist()
        )
