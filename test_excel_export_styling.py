from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from utils.export_helpers import (
    COLUMN_WIDTH_MARGIN,
    EXPORT_TYPE_COMBINED,
    EXPORT_TYPE_PER_CALCULATION,
    EXPORT_TYPE_SINGLE,
    FIREBOARD_ANCHOR_COLUMN,
    FIREBOARD_ANCHOR_ROW,
    FIREBOARD_GREY,
    FIREBOARD_TEXT,
    KNAUF_BLUE,
    PRINT_HEADER_FOOTER_MARGIN_INCHES,
    PRINT_MARGIN_INCHES,
    TABLE_START_ROW,
    add_system_separator_rows,
    create_materials_excel,
    get_material_list_title,
    get_materials_excel_filename,
)


MATERIAL_COLUMNS_DA = [
    "ART.NR.",
    "DB.nr.",
    "Producent",
    "Materiale",
    "Forbrug",
    "Enhed",
    "Total",
]


def _sample_materials_df():
    return pd.DataFrame(
        [
            ["2906", "5959671", "Knauf A/S", "15 mm Fireboard", 1.5, "m²", 3.0],
            ["181533", "8156069", "Knauf A/S", "Skrue RAB25", 2.0, "stk", 20.0],
        ],
        columns=MATERIAL_COLUMNS_DA,
    )


def _load_workbook_from_export(excel_file):
    excel_file.seek(0)
    return load_workbook(BytesIO(excel_file.read()))


def _assert_header_style(row):
    for cell in row:
        assert cell.fill.fgColor.rgb == f"00{KNAUF_BLUE}"
        assert cell.font.color.rgb == "00FFFFFF"
        assert cell.font.bold is True


def _assert_fireboard_brand_text(worksheet):
    fireboard_cell = worksheet.cell(
        row=FIREBOARD_ANCHOR_ROW,
        column=FIREBOARD_ANCHOR_COLUMN,
    )
    assert fireboard_cell.value == FIREBOARD_TEXT
    assert fireboard_cell.font.color.rgb == f"00{FIREBOARD_GREY}"
    assert fireboard_cell.font.italic is True


def _assert_print_layout(worksheet, header_row):
    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.fitToWidth == 1
    assert worksheet.page_setup.fitToHeight == 0
    assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert worksheet.print_options.horizontalCentered is True
    assert worksheet.print_title_rows == f"${header_row}:${header_row}"
    assert worksheet.page_margins.left == PRINT_MARGIN_INCHES
    assert worksheet.page_margins.right == PRINT_MARGIN_INCHES
    assert worksheet.page_margins.top == PRINT_MARGIN_INCHES
    assert worksheet.page_margins.bottom == PRINT_MARGIN_INCHES
    assert worksheet.page_margins.header == PRINT_HEADER_FOOTER_MARGIN_INCHES
    assert worksheet.page_margins.footer == PRINT_HEADER_FOOTER_MARGIN_INCHES


def _expected_visible_column_width(worksheet, column_number):
    max_length = 0

    for row_number in range(1, worksheet.max_row + 1):
        if worksheet.row_dimensions[row_number].hidden:
            continue

        value = worksheet.cell(row=row_number, column=column_number).value

        if value is not None:
            max_length = max(max_length, len(str(value)))

    return max_length + COLUMN_WIDTH_MARGIN


def _assert_columns_autosized_to_visible_values(worksheet):
    for column_number in range(1, worksheet.max_column + 1):
        column_letter = worksheet.cell(
            row=1,
            column=column_number
        ).column_letter
        assert worksheet.column_dimensions[column_letter].width == (
            _expected_visible_column_width(worksheet, column_number)
        )


def _worksheet_values(worksheet, min_row, max_row, max_col):
    return [
        [
            worksheet.cell(row=row, column=column).value
            for column in range(1, max_col + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]


def test_normal_material_list_excel_is_styled_without_changing_data_values():
    source_df = _sample_materials_df()

    workbook = _load_workbook_from_export(
        create_materials_excel(source_df, language="DA")
    )
    worksheet = workbook["Materialeliste"]

    assert worksheet.cell(row=4, column=1).value == "Materialeliste"
    assert len(worksheet._images) == 1
    _assert_fireboard_brand_text(worksheet)
    _assert_header_style(worksheet[TABLE_START_ROW])
    assert worksheet.freeze_panes == f"A{TABLE_START_ROW + 1}"
    _assert_columns_autosized_to_visible_values(worksheet)
    _assert_print_layout(worksheet, TABLE_START_ROW)

    exported_values = _worksheet_values(
        worksheet,
        TABLE_START_ROW + 1,
        TABLE_START_ROW + len(source_df),
        len(source_df.columns),
    )
    assert exported_values == source_df.values.tolist()


def test_combined_material_list_excel_is_styled_without_changing_aggregation_values():
    aggregated_df = pd.DataFrame(
        [
            ["2906", "5959671", "Knauf A/S", "15 mm Fireboard", 3.0, "m²", 6.0],
            ["181533", "8156069", "Knauf A/S", "Screw RAB25", 4.0, "pcs", 40.0],
        ],
        columns=[
            "ART.NO.",
            "DB no.",
            "Manufacturer",
            "Material",
            "Consumption",
            "Unit",
            "Total",
        ],
    )

    workbook = _load_workbook_from_export(
        create_materials_excel(
            aggregated_df,
            language="EN",
            export_type=EXPORT_TYPE_COMBINED,
        )
    )
    worksheet = workbook["Materialeliste"]

    assert worksheet.cell(row=4, column=1).value == "Combined Material List"
    assert len(worksheet._images) == 1
    _assert_fireboard_brand_text(worksheet)
    _assert_header_style(worksheet[TABLE_START_ROW])
    _assert_columns_autosized_to_visible_values(worksheet)
    _assert_print_layout(worksheet, TABLE_START_ROW)

    exported_values = _worksheet_values(
        worksheet,
        TABLE_START_ROW + 1,
        TABLE_START_ROW + len(aggregated_df),
        len(aggregated_df.columns),
    )
    assert exported_values == aggregated_df.values.tolist()


def test_column_autosizing_adds_enough_padding_for_long_headers():
    danish_df = pd.DataFrame(
        [["2906", 3.0]],
        columns=["ART.NR.", "SAMLET MÆNGDE"],
    )
    english_df = pd.DataFrame(
        [["1", "Knauf A/S"]],
        columns=["Art", "MANUFACTURER"],
    )

    danish_workbook = _load_workbook_from_export(
        create_materials_excel(danish_df, language="DA")
    )
    english_workbook = _load_workbook_from_export(
        create_materials_excel(english_df, language="EN")
    )

    danish_worksheet = danish_workbook["Materialeliste"]
    english_worksheet = english_workbook["Materialeliste"]

    _assert_columns_autosized_to_visible_values(danish_worksheet)
    _assert_columns_autosized_to_visible_values(english_worksheet)
    assert danish_worksheet.column_dimensions["B"].width == (
        len("SAMLET MÆNGDE") + COLUMN_WIDTH_MARGIN
    )
    assert english_worksheet.column_dimensions["B"].width == (
        len("MANUFACTURER") + COLUMN_WIDTH_MARGIN
    )


def test_per_system_material_list_preserves_structure_and_styles_repeated_headers():
    materials_df = pd.DataFrame(
        [
            ["System B", "181533", "Skrue RAB25", 2.0],
            ["System A", "2906", "15 mm Fireboard", 1.5],
            ["System B", "181534", "Skrue RAB35", 4.0],
        ],
        columns=["SYSTEM", "ART.NR.", "Materiale", "Total"],
    )
    per_system_df = add_system_separator_rows(materials_df)

    workbook = _load_workbook_from_export(
        create_materials_excel(
            per_system_df,
            include_header=False,
            language="DA",
            per_system=True,
        )
    )
    worksheet = workbook["Materialeliste"]

    assert worksheet.cell(row=4, column=1).value == "Materialeliste pr. beregning"
    assert len(worksheet._images) == 1
    _assert_fireboard_brand_text(worksheet)
    assert worksheet.cell(row=TABLE_START_ROW, column=1).value == "SYSTEM"
    assert worksheet.freeze_panes == f"A{TABLE_START_ROW + 1}"
    _assert_columns_autosized_to_visible_values(worksheet)
    _assert_print_layout(worksheet, TABLE_START_ROW)

    repeated_header_rows = []
    for row_number in range(TABLE_START_ROW, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_number, column=column).value
            for column in range(1, len(per_system_df.columns) + 1)
        ]
        if row_values == list(per_system_df.columns):
            repeated_header_rows.append(row_number)
            _assert_header_style(worksheet[row_number])

    assert len(repeated_header_rows) == 2

    system_name_rows = [TABLE_START_ROW + 1, TABLE_START_ROW + 5]
    for row_number in system_name_rows:
        assert worksheet.cell(row=row_number, column=1).font.bold is True
        assert worksheet.cell(row=row_number, column=1).fill.fgColor.rgb in (
            "00000000",
            "000000",
        )

    exported_values = _worksheet_values(
        worksheet,
        TABLE_START_ROW,
        TABLE_START_ROW + len(per_system_df) - 1,
        len(per_system_df.columns),
    )
    expected_values = per_system_df.where(pd.notna(per_system_df), None).values.tolist()
    expected_values = [
        [None if value == "" else value for value in row]
        for row in expected_values
    ]
    assert exported_values == expected_values


def test_material_list_workbook_titles_are_localized_for_all_export_types():
    materials_df = _sample_materials_df()

    scenarios = [
        ("DA", EXPORT_TYPE_SINGLE, False, True, "Materialeliste"),
        ("DA", EXPORT_TYPE_COMBINED, False, True, "Samlet materialeliste"),
        (
            "DA",
            EXPORT_TYPE_PER_CALCULATION,
            True,
            False,
            "Materialeliste pr. beregning",
        ),
        ("EN", EXPORT_TYPE_SINGLE, False, True, "Material List"),
        ("EN", EXPORT_TYPE_COMBINED, False, True, "Combined Material List"),
        (
            "EN",
            EXPORT_TYPE_PER_CALCULATION,
            True,
            False,
            "Material List per Calculation",
        ),
    ]

    for language, export_type, per_system, include_header, expected_title in scenarios:
        export_df = (
            add_system_separator_rows(materials_df)
            if per_system
            else materials_df
        )
        workbook = _load_workbook_from_export(
            create_materials_excel(
                export_df,
                include_header=include_header,
                language=language,
                per_system=per_system,
                export_type=export_type,
            )
        )
        worksheet = workbook["Materialeliste"]

        assert worksheet.cell(row=4, column=1).value == expected_title


def test_material_list_titles_are_localized_for_all_excel_export_types():
    assert get_material_list_title("DA", EXPORT_TYPE_SINGLE) == "Materialeliste"
    assert get_material_list_title("DA", EXPORT_TYPE_COMBINED) == (
        "Samlet materialeliste"
    )
    assert get_material_list_title("DA", EXPORT_TYPE_PER_CALCULATION) == (
        "Materialeliste pr. beregning"
    )
    assert get_material_list_title("EN", EXPORT_TYPE_SINGLE) == "Material List"
    assert get_material_list_title("EN", EXPORT_TYPE_COMBINED) == (
        "Combined Material List"
    )
    assert get_material_list_title("EN", EXPORT_TYPE_PER_CALCULATION) == (
        "Material List per Calculation"
    )


def test_material_list_download_filenames_are_localized_for_all_export_types():
    assert get_materials_excel_filename("DA", EXPORT_TYPE_SINGLE) == "Materialeliste.xlsx"
    assert get_materials_excel_filename("DA", EXPORT_TYPE_COMBINED) == (
        "Samlet_materialeliste.xlsx"
    )
    assert get_materials_excel_filename("DA", EXPORT_TYPE_PER_CALCULATION) == (
        "Materialeliste_pr_beregning.xlsx"
    )
    assert get_materials_excel_filename("EN", EXPORT_TYPE_SINGLE) == "Material_List.xlsx"
    assert get_materials_excel_filename("EN", EXPORT_TYPE_COMBINED) == (
        "Combined_Material_List.xlsx"
    )
    assert get_materials_excel_filename("EN", EXPORT_TYPE_PER_CALCULATION) == (
        "Material_List_per_Calculation.xlsx"
    )
